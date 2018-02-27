# /usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from openpyxl import Workbook
import datetime

wb = load_workbook('courses.xlsx')
student_sheet =  wb['students']
time_sheet = wb['time']


def combine():
    combine_sheet = wb.create_sheet('combine')
    combine_sheet.append(['创建时间', '课程名称', '学习人数', '学习时间'])

    for stu in student_sheet.values:
        if stu[2] != '学习人数':
            for time in time_sheet.values:
                if stu[1] == time[1]:
                    combine_sheet.append(list(stu) + [time[2]])

    wb.save('courses.xlsx')


def split():
    combine_sheet = wb['combine']
    split_year = []
    for item in combine_sheet.values:
        if item[2] != '学习人数':
            split_year.append(item[0].strftime('%Y'))

    for year in set(split_year):
        wb_tmp = Workbook()
        wb_tmp.remove(wb_tmp.active)
        new_sheet = wb_tmp.create_sheet(year)

        for item in combine_sheet.values:
            if item[2] != '学习人数':
                if item[0].strftime('%Y') == year:
                    new_sheet.append(item)
        wb_tmp.save('{}.xlsx'.format(year))


if __name__ == '__main__':
    combine()
    split()